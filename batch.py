import openpyxl
import unicodedata
import re
#fichier de données contenant la liste des personnels avec leurs numens
#RGPD -> attention à la conservation de ce fichier 
from data import numens as numens

def nettoyer_fichier(valeur):
    #les valeurs saisies dans le fichier liste doit être nettoyées et normalisées pour pouvoir être comparées
    if valeur is not None:
        #encodage du fichier
        valeur=unicodedata.normalize('NFKD',valeur).encode('ASCII', 'ignore').decode('ASCII')
        #suppression des espaces avant et après la valeur
        valeur=valeur.strip()
        #passage de toutes les valeurs en caractères minuscules
        valeur=valeur.upper()
        #remplacement des - en espace
        if valeur.find('-'):
            valeur=valeur.replace('-',' ')
    return valeur

#ouverture du fichier liste
wb = openpyxl.load_workbook('liste.xlsx')
#référencement de l'onglet actif 
onglet = wb.active


liste_titre_colonne = []
personnes = []

#écrire les résultats dans un fichier ficCandidat.unl
ficCandidatFile = open('ficCandidat.unl', 'w')
#écrire les erreurs (personnes non trouvées et/ou personnes avec plusieurs identifications possibles
erreurFile = open('erreur.txt','w')
#écrire la liste des disciplines pour vérification (très important) pour éviter les erreurs d'identification
verificationFile = open('verification.txt','w')

#récupération des titres des colonnes du fichier liste
for i in range(1,onglet.max_column+1):
    liste_titre_colonne.append(onglet.cell(row=1,column=i).value)
#récupération de la liste des personnes à convoquer
for i in range(2,onglet.max_row+1):
    personne = {}
    for j in range(0,onglet.max_column):
        personne[liste_titre_colonne[j]]=onglet.cell(row=i,column=j+1).value
    personnes.append(personne)


#nettoyer le fichier de données
for ligne in personnes:
    for cle in ligne.keys():
        if cle in ('Nom','Prenom','Discipline','Ville','RNE'):
            ligne[cle]=nettoyer_fichier(ligne[cle])


global trouve
trouve=0
#résultat si tous les champs de la liste de personnes à convoquer concordent avec la liste des numens
resultat_fort = []
##résultat si certains champs de la liste de personnes à convoquer concordent avec la liste des numens
resultat_faible = []

#chercher les personnes   
for ligne in personnes:

    #selon deux listes resultats faibles ou forts
    for numen in numens:
        if (numen['Nom usage'] is not None and numen['Nom patronymique'] is not None) and numen['Prenom'] is not None:
            if (ligne['Nom'] == numen['Nom usage'] or ligne['Nom'] == numen['Nom patronymique']) and  ligne['Prenom'] == numen['Prenom']:
                if numen['Discipline'] is not None and ligne['Discipline'] is not None :
                    if ligne['Discipline'] in numen['Discipline']:
                        resultat_fort.append(numen)
                if ligne['RNE'] is not None and numen['RNE'] is not None:
                    if ligne['RNE'] == numen['RNE']and numen not in resultat_fort:
                        resultat_fort.append(numen)
                if ligne['Ville'] is not None and numen['Ville'] is not None:
                    if ligne['Ville'] in numen['Ville'] and numen not in resultat_fort:
                        resultat_fort.append(numen)
                if len(resultat_fort)==0:
                    resultat_faible.append(numen)

            if (ligne['Nom'] in numen['Nom usage'] or ligne['Nom'] in numen['Nom patronymique'] or numen['Nom usage'] in ligne['Nom'] or numen['Nom patronymique'] in ligne['Nom'] ) and  (ligne['Prenom'] in numen['Prenom'] or numen['Prenom'] in ligne['Prenom']):
                if numen['Discipline'] is not None and ligne['Discipline'] is not None :
                    if ligne['Discipline'] in numen['Discipline']:
                        resultat_faible.append(numen)
                if ligne['RNE'] is not None and numen['RNE'] is not None:
                    if ligne['RNE'] == numen['RNE']and numen not in resultat_faible:
                        resultat_faible.append(numen)
                if ligne['Ville'] is not None and numen['Ville'] is not None:
                    if ligne['Ville'] in numen['Ville'] and numen not in resultat_faible:
                        resultat_faible.append(numen)
                if len(resultat_fort)==0 and len(resultat_faible)==0:
                    resultat_faible.append(numen)
                    
    #écriture des résultats dans les fichiers erreurs, vérification et ficCandidat.unl
    if len(resultat_fort)==1:
        for numen in resultat_fort:
            candidature = numen['Numen'] + '|' + str(ligne['Dispositif']) + '|' + str(ligne['Module']) + '|' + str(ligne['Groupe']).rjust(2,'0') + '|R|\n'
            if numen['Discipline'] is None:
                numen['Discipline'] = 'sans specialite'
            verification = numen['Numen'] + ' ' + numen['Nom usage'] + ' ' + ' ' + numen['Prenom'] + ' ' +  numen['Discipline'] + '\n'
            verificationFile.write(verification)
            ficCandidatFile.write(candidature)

    if len(resultat_faible)==1 and len(resultat_fort)==0:
        for numen in resultat_faible:
            candidature = numen['Numen'] + '|' + str(ligne['Dispositif']) + '|' + str(ligne['Module']) + '|' + str(ligne['Groupe']).rjust(2,'0') + '|R|\n'
            if numen['Discipline'] is None:
                numen['Discipline'] = 'sans specialite'
            verification = numen['Numen'] + ' ' + numen['Nom usage'] + ' ' + ' ' + numen['Prenom'] + ' ' +  numen['Discipline'] + '\n'
            verificationFile.write(verification)
            ficCandidatFile.write(candidature)

    if len(resultat_fort)>1 :
        erreurFile.write('Plusieurs candidats possibles pour ' + ligne['Nom'] + ' ' + ligne['Prenom'] + '\n')

        for numen in resultat_fort:
            candidature = numen['Numen'] + '|' + str(ligne['Dispositif']) + '|'
            candidature = candidature + str(ligne['Module']) + '|'
            if numen['Discipline'] is None:
                numen['Discipline'] = 'sans specialite'
            candidature = candidature + str(ligne['Groupe']).rjust(2,'0') + '|R|' + ' ' + numen['Nom usage'] + ' ' + numen['Prenom'] + ' ' + numen['Discipline'] + ' ' + numen['RNE'] + ' ' + numen['Ville'] + '\n'
            erreurFile.write(candidature)

    if len(resultat_faible)>1 and len(resultat_fort)==0:
        erreurFile.write('Plusieurs candidats possibles pour ' + ligne['Nom'] + ' ' + ligne['Prenom'] + '\n')

        for numen in resultat_faible:
            candidature = numen['Numen'] + '|' + str(ligne['Dispositif']) + '|'
            candidature = candidature + str(ligne['Module']) + '|'
            if numen['Discipline'] is None:
                numen['Discipline'] = 'sans specialite'
            candidature = candidature + str(ligne['Groupe']).rjust(2,'0') + '|R|' + ' ' + numen['Nom usage'] + ' ' + numen['Prenom'] + ' ' + numen['Discipline'] + ' ' + numen['RNE'] + ' ' + numen['Ville'] + '\n'
            erreurFile.write(candidature)

    if len(resultat_fort)==0 and len(resultat_faible)==0:
        erreurFile.write('Candidat non trouve \n')
        candidature = ligne['Nom'] + ' ' + ligne['Prenom']
        if ligne ['Discipline'] is None :
            ligne['Discipline'] = 'sans specialite'
        candidature = candidature + ' ' + ligne ['Discipline']
        if ligne['RNE'] is  None:
            ligne['RNE'] = 'Sans RNE'
        candidature = candidature    + ' ' + ligne['RNE']
        if ligne['Ville'] is None:
            ligne['Ville'] = 'sans ville'
        candidature = candidature + ' ' + ligne['Ville']
        candidature = candidature + '\n'
        erreurFile.write(candidature)

    resultat_fort = []
    resultat_faible = []




ficCandidatFile.close()
erreurFile.close()
verificationFile.close()
